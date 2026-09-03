"""
Generate the Excel workbooks that back the dashboard - one per chapter.

Headers here must stay identical to src/data/columns.ts. Row values are seed
examples; once the sheets are connected, the Excel is the source of truth.

Run:  python scripts/generate_excel.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "excel"

FONT = "Arial"
MONEY = '"₹"#,##0'
PERCENT = '0"%"'          # stored 0-100 to match the app's completion field
COUNT = "#,##0"
YEAR = "0"

INPUT_FONT = Font(name=FONT, size=10, color="0000FF")      # typed by hand
CALC_FONT = Font(name=FONT, size=10, color="000000")       # formula
LOCKED_FONT = Font(name=FONT, size=10, color="9C9C9C")     # do not edit
BODY_FONT = Font(name=FONT, size=10)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

PROJECT_HEADERS = [
    ("Live projects", 22, None),
    ("Percentage of completion (Lindsay)", 16, PERCENT),
    ("Total sold assets (CRM)", 14, COUNT),
    ("Inventory (Lincoln)", 12, COUNT),
    ("Expected outcome (Lindsay)", 18, MONEY),
    ("Expected handover (Lindsay)", 16, None),
    ("Amount repaid to first project (Lindsay)", 20, MONEY),
    ("Year started (CRM)", 12, YEAR),
    ("Comments (Lincoln)", 46, None),
]

PIPELINE_HEADERS = [
    ("Pipeline project", 22, None),
    ("Onward (Lincoln)", 14, None),
    ("Inventory (Lincoln/Lindsay)", 16, COUNT),
    ("Ticket size (Lincoln)", 16, MONEY),
    ("Total cost of construction (Lincoln/Lindsay)", 22, MONEY),
    ("Planned (Lincoln)", 12, COUNT),
    ("Status (Lincoln)", 22, None),
]

# No owner brackets: the lenders sheet is maintained jointly, and an owner
# here would put the wrong name on every entry in the dashboard's update strip.
LIABILITY_HEADERS = [
    ("Lender", 26, None),
    ("Loan amount", 18, MONEY),
    ("Outstanding", 18, MONEY),
]

OPEX_HEADERS = [
    ("Opex", 22, None),
    ("Amount opex (Lincoln)", 16, MONEY),
    ("Extra (Lincoln)", 14, MONEY),
    ("Remark (Lincoln)", 40, None),
    ("Allocated amount + extra", 20, MONEY),
]

CHAPTERS = [
    {
        "id": "goa",
        "name": "Goa Chapter",
        "accent": "amber",
        "timeZone": "Asia/Kolkata",
        "timeZoneLabel": "Goa, India",
        "headerFill": "FDF3D8",
        "headerColor": "7A4A05",
        "crmYear": 2026,
        "projects": [
            ["goa-p1", "Project One", 100, 24, 0, 180000000, "Delivered", 0, 2019,
             "Fully handed over, closing formalities done."],
            ["goa-p2", "Project Two", 82, 18, 6, 240000000, "Mar 2026", 45000000, 2021,
             "Finishing works on track; 6 units still open."],
            ["goa-p3", "Project Three", 54, 11, 14, 310000000, "Dec 2026", 20000000, 2023,
             "Structure complete, MEP in progress."],
            ["goa-p4", "Project Four", 18, 4, 22, 400000000, "Q3 2027", 0, 2025,
             "Foundation stage; launch pricing under review."],
        ],
        "pipeline": [
            ["goa-pl1", "Pipeline One", "Apr 2026", 30, 25000000, 450000000, 30, "Land acquired"],
            ["goa-pl2", "Pipeline Two", "Oct 2026", 18, 32000000, 380000000, 18, "Approvals in progress"],
            ["goa-pl3", "Pipeline Three", "Q2 2027", 42, 18000000, 520000000, 48, "Design stage"],
        ],
        "opex": [
            ["goa-o1", "Site office", 1200000, 150000, "Lease renewed for the year."],
            ["goa-o2", "Marketing", 3500000, 400000, "Extra spend on the launch campaign."],
            ["goa-o3", "Salaries", 8400000, 0, ""],
            ["goa-o4", "Maintenance", 950000, 75000, "AMC revised in April."],
        ],
        "crm": {
            "goa-p1": {"Jan": 2000000, "Feb": 1500000, "Mar": 600000},
            "goa-p2": {"Jan": 7200000, "Feb": 6800000, "Mar": 8100000, "Apr": 5400000,
                       "May": 7900000, "Jun": 6200000, "Jul": 8600000, "Aug": 7100000},
            "goa-p3": {"Jan": 3400000, "Feb": 4100000, "Mar": 3900000, "Apr": 5200000,
                       "May": 4600000, "Jun": 5800000, "Jul": 4300000, "Aug": 6000000},
            "goa-p4": {"Apr": 1200000, "May": 2400000, "Jun": 1900000, "Jul": 3100000,
                       "Aug": 2700000},
        },
    },
    {
        "id": "portugal",
        "name": "Portugal Chapter",
        "accent": "indigo",
        "timeZone": "Europe/Lisbon",
        "timeZoneLabel": "Portugal",
        "headerFill": "E6E7FB",
        "headerColor": "2F2C7A",
        "crmYear": 2026,
        "projects": [
            ["pt-p1", "Project One", 91, 14, 2, 120000000, "Jan 2026", 30000000, 2022,
             "Snagging under way, handover set for January."],
            ["pt-p2", "Project Two", 47, 9, 12, 210000000, "Nov 2026", 15000000, 2024,
             "Facade works started; sales pace steady."],
            ["pt-p3", "Project Three", 12, 3, 19, 260000000, "Q4 2027", 0, 2025,
             "Excavation stage."],
        ],
        "pipeline": [
            ["pt-pl1", "Pipeline One", "Jun 2026", 22, 28000000, 300000000, 22, "Land acquired"],
            ["pt-pl2", "Pipeline Two", "Q1 2027", 16, 35000000, 340000000, 20, "Approvals in progress"],
        ],
        "opex": [
            ["pt-o1", "Site office", 800000, 90000, "Shared with Project Two."],
            ["pt-o2", "Marketing", 2100000, 260000, "Portal listings added."],
            ["pt-o3", "Salaries", 5600000, 0, ""],
            ["pt-o4", "Maintenance", 620000, 40000, "Landscaping contract."],
        ],
        "crm": {
            "pt-p1": {"Jan": 4800000, "Feb": 5200000, "Mar": 4100000, "Apr": 3600000,
                      "May": 2900000, "Jun": 2200000},
            "pt-p2": {"Jan": 3100000, "Feb": 2800000, "Mar": 3700000, "Apr": 4200000,
                      "May": 3900000, "Jun": 4500000, "Jul": 3300000, "Aug": 4800000},
            "pt-p3": {"Jun": 900000, "Jul": 1600000, "Aug": 2100000},
        },
    },
]


def write_header(ws, headers, chapter):
    """Header row and column widths.

    The id column is appended last so the data columns keep the same letters as
    the dashboard - on the Opex sheet, E is literally B plus C.
    """
    fill = PatternFill("solid", fgColor=chapter["headerFill"])
    font = Font(name=FONT, size=10, bold=True, color=chapter["headerColor"])

    cols = list(headers) + [("id (do not edit)", 14, None)]

    for i, (label, width, _fmt) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="bottom", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def write_rows(ws, headers, rows, chapter):
    """Rows arrive as [id, ...values]; the id goes in the trailing column."""
    write_header(ws, headers, chapter)
    for r, row in enumerate(rows, start=2):
        row_id, values = row[0], row[1:]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = INPUT_FONT
            number_format = headers[c - 1][2]
            if number_format:
                cell.number_format = number_format
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=len(headers) + 1, value=row_id).font = LOCKED_FONT


def build_opex(ws, chapter):
    write_header(ws, OPEX_HEADERS, chapter)
    for r, (row_id, name, amount, extra, remark) in enumerate(chapter["opex"], start=2):
        ws.cell(row=r, column=1, value=name).font = INPUT_FONT

        amount_cell = ws.cell(row=r, column=2, value=amount)
        amount_cell.font = INPUT_FONT
        amount_cell.number_format = MONEY

        extra_cell = ws.cell(row=r, column=3, value=extra)
        extra_cell.font = INPUT_FONT
        extra_cell.number_format = MONEY

        remark_cell = ws.cell(row=r, column=4, value=remark)
        remark_cell.font = INPUT_FONT
        remark_cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Calculated column, exactly as specified: column B plus column C.
        calc = ws.cell(row=r, column=5, value="=B{0}+C{0}".format(r))
        calc.font = CALC_FONT
        calc.number_format = MONEY

        ws.cell(row=r, column=6, value=row_id).font = LOCKED_FONT


def build_crm(ws, chapter):
    fill = PatternFill("solid", fgColor=chapter["headerFill"])
    head_font = Font(name=FONT, size=10, bold=True, color=chapter["headerColor"])

    projects = chapter["projects"]

    month_head = ws.cell(row=1, column=1, value="Month (Lendl)")
    month_head.fill = fill
    month_head.font = head_font
    ws.column_dimensions["A"].width = 14

    for i, project in enumerate(projects, start=2):
        cell = ws.cell(row=1, column=i, value=project[1])
        cell.fill = fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="bottom", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "B2"

    for r, month in enumerate(MONTHS, start=2):
        ws.cell(row=r, column=1, value=month).font = BODY_FONT
        for i, project in enumerate(projects, start=2):
            amount = chapter["crm"].get(project[0], {}).get(month)
            cell = ws.cell(row=r, column=i, value=amount)
            cell.font = INPUT_FONT
            cell.number_format = MONEY

    total_row = len(MONTHS) + 2
    label = ws.cell(row=total_row, column=1, value="Total")
    label.font = Font(name=FONT, size=10, bold=True)
    label.fill = fill

    for i in range(2, len(projects) + 2):
        letter = get_column_letter(i)
        cell = ws.cell(row=total_row, column=i,
                       value="=SUM({0}2:{0}{1})".format(letter, total_row - 1))
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.number_format = MONEY
        cell.fill = fill


def build_legend(ws, chapter):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 64

    title = ws.cell(row=1, column=1, value=chapter["name"])
    title.font = Font(name=FONT, size=14, bold=True, color=chapter["headerColor"])

    lines = [
        ("", ""),
        ("Chapter settings", ""),
        ("id", chapter["id"]),
        ("name", chapter["name"]),
        ("accent", chapter["accent"]),
        ("timeZone", chapter["timeZone"]),
        ("timeZoneLabel", chapter["timeZoneLabel"]),
        ("CRM Collection year", chapter["crmYear"]),
        ("", ""),
        ("How to use this workbook", ""),
        ("Blue cells", "Type in these. Everything you maintain is blue."),
        ("Black cells", "Formulas. Do not overwrite - they recalculate."),
        ("Grey id column", "The join key. Never edit or reorder it."),
        ("", "CRM Collection maps to Live Projects through it."),
        ("", ""),
        ("Formats", ""),
        ("Money", "Whole rupees, e.g. 240000000. The dashboard shows ₹24Cr."),
        ("Percentage of completion", "A number from 0 to 100, e.g. 82. Not 0.82."),
        ("Expected handover / Onward", "Free text, e.g. Mar 2026 or Q3 2027."),
        ("Handover value Delivered", "Spelled exactly, it shows as a green pill."),
        ("Status", "Land acquired / Approvals in progress / Design stage /"),
        ("", "Yard slot booked. Anything else works, it just shows grey."),
        ("Empty CRM month", "Leave blank. The dashboard shows a dash."),
        ("", ""),
        ("Rows", "Row 2 down are real examples - overwrite them or add below."),
        ("Adding a project", "Add the row in Live Projects, then add its column"),
        ("", "in CRM Collection with the same project name."),
        ("Allocated amount + extra", "Calculated as Amount opex plus Extra (column B + C)."),
        ("", ""),
        ("Source of the seed values", "Placeholder figures supplied by the dashboard build,"),
        ("", "not real accounts. Replace them all."),
    ]

    for offset, (key, value) in enumerate(lines, start=2):
        key_cell = ws.cell(row=offset, column=1, value=key)
        key_cell.font = Font(name=FONT, size=10, bold=bool(key) and not value)
        value_cell = ws.cell(row=offset, column=2, value=value)
        value_cell.font = BODY_FONT
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")


def build(chapter):
    wb = Workbook()

    legend = wb.active
    legend.title = "Chapter"
    build_legend(legend, chapter)

    write_rows(wb.create_sheet("Live Projects"), PROJECT_HEADERS,
               chapter["projects"], chapter)
    write_rows(wb.create_sheet("Pipeline"), PIPELINE_HEADERS,
               chapter["pipeline"], chapter)
    build_opex(wb.create_sheet("Opex"), chapter)
    # Headers only - the lenders are filled in by hand, and inventing seed rows
    # here would put fictional debts on the board.
    write_rows(wb.create_sheet("Liabilities"), LIABILITY_HEADERS,
               chapter.get("liabilities", []), chapter)
    build_crm(wb.create_sheet("CRM Collection"), chapter)

    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / (chapter["name"] + ".xlsx")
    wb.save(path)
    print("wrote", path)
    return path


if __name__ == "__main__":
    for chapter_def in CHAPTERS:
        build(chapter_def)
